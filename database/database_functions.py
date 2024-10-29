from openpyxl import load_workbook
import pandas as pd
import requests
import os

def adicionar_linha_excel(nome, numero, msg_usuario, msg_gpt):
    caminho_arquivo = os.path.join("database", "registro.xlsx")

    if not os.path.exists(caminho_arquivo):
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["nome", "numero", "msg_usuario", "msg_gpt", "hora"])  # Cabeçalhos
        workbook.save(caminho_arquivo)

    workbook = load_workbook(caminho_arquivo)
    sheet = workbook.active
    sheet.append([nome, numero, msg_usuario, msg_gpt])
    workbook.save(caminho_arquivo)

def visualizar_registros_excel():
    caminho_arquivo = os.path.join(os.path.dirname(__file__), "registro.xlsx")    
    
    try:
        # Lê o arquivo Excel usando pandas
        df = pd.read_excel(caminho_arquivo)
        
        # Converte o DataFrame para uma lista de dicionários
        registros = df.to_dict(orient="records")
        
        return registros

    except FileNotFoundError:
        return {"error": "Arquivo não encontrado"}
    except Exception as e:
        return {"error": str(e)}
    
def enviar_mensagem_whatsapp(numero, mensagem):
    url = "https://api.z-api.io/instances/3D699FAFFEADD094C8E42E5479B6AFF4/token/6797E7BEE32128FFAD4EEF61/send-text"
    headers = {
        "Content-Type": "application/json",
        "Client-Token": "F885b84cd15ed441da1a4395a2aafea14S"
    }
    payload = {
        "phone": numero,
        "message": mensagem
    }

    response = requests.post(url, json=payload, headers=headers)
    return response.json()